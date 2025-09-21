import hashlib
import json
import time
from datetime import timedelta
import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import TemplateView, FormView, DetailView

from apps.core.services.ai_service import AIService
from .forms import ExcelUploadForm
from .models import ExcelUpload, ExcelData, AIValidation


class ExcelManagerView(LoginRequiredMixin, TemplateView):
    """Main page with upload area and file list."""

    template_name = "excel_manager/index.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Excel Manager"
        context["form"] = ExcelUploadForm()
        context["uploads"] = ExcelUpload.objects.filter(
            user=self.request.user
        ).select_related("user")[:10]
        return context


class ExcelUploadView(LoginRequiredMixin, FormView):
    """HTMX-aware view for Excel upload."""

    form_class = ExcelUploadForm
    template_name = "excel_manager/partials/_upload_area.html"
    success_url = reverse_lazy("excel_manager:index")

    def form_valid(self, form):
        """Handle successful form submission."""
        excel_file = form.cleaned_data["file"]

        # Calculate file hash for duplicate detection
        sha256_hash = hashlib.sha256()
        for chunk in excel_file.chunks():
            sha256_hash.update(chunk)
        file_hash = sha256_hash.hexdigest()
        excel_file.seek(0)

        # Check for duplicates
        existing_upload = ExcelUpload.objects.filter(
            file_hash=file_hash, user=self.request.user
        ).first()

        if existing_upload:
            # File already uploaded by this user
            if hasattr(self.request, "htmx") and self.request.htmx:
                context = {
                    "error": f"This file was already uploaded on {existing_upload.uploaded_at}",
                    "form": form,
                }
                return render(
                    self.request,
                    "excel_manager/partials/_upload_area.html",
                    context,
                    status=400,
                )
            messages.warning(self.request, "This file has already been uploaded.")
            return super().form_valid(form)

        try:
            with transaction.atomic():
                # Create ExcelUpload record
                upload = ExcelUpload.objects.create(
                    user=self.request.user,
                    file=excel_file,
                    original_filename=excel_file.name,
                    file_hash=file_hash,
                    file_size=excel_file.size,
                    status=ExcelUpload.STATUS_PROCESSING,
                )

                # Process Excel file synchronously (MVP)
                self.process_excel_file(upload)

                # Mark as completed
                upload.status = ExcelUpload.STATUS_COMPLETED
                from django.utils import timezone

                upload.processed_at = timezone.now()
                upload.save()

            # Return success response for HTMX
            if hasattr(self.request, "htmx") and self.request.htmx:
                context = {
                    "uploads": ExcelUpload.objects.filter(
                        user=self.request.user
                    ).select_related("user")[:10],
                    "success_message": "Excel file uploaded successfully!",
                }
                response = render(
                    self.request, "excel_manager/partials/_file_list.html", context
                )
                # Trigger event to reset upload form
                response["HX-Trigger"] = "excel-uploaded"
                return response

            messages.success(self.request, "Excel file uploaded successfully!")
            return super().form_valid(form)

        except Exception as e:
            # Mark as failed
            if upload:
                upload.status = ExcelUpload.STATUS_FAILED
                upload.error_message = str(e)
                upload.save()

            # Return error response
            if hasattr(self.request, "htmx") and self.request.htmx:
                context = {
                    "error": f"Failed to process file: {str(e)}",
                    "form": form,
                }
                return render(
                    self.request,
                    "excel_manager/partials/_upload_area.html",
                    context,
                    status=400,
                )

            messages.error(self.request, f"Failed to process file: {str(e)}")
            return self.form_invalid(form)

    def process_excel_file(self, upload):
        """Process the Excel file and store data."""
        try:
            # Load workbook
            wb = openpyxl.load_workbook(upload.file, data_only=True)
            sheet_count = len(wb.sheetnames)
            upload.sheet_count = sheet_count
            upload.save()

            # Process each sheet
            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]

                # Get headers from first row
                headers = []
                first_row = next(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
                )
                if first_row:
                    headers = [
                        str(cell) if cell is not None else f"Column {i+1}"
                        for i, cell in enumerate(first_row)
                    ]

                # Get data rows (limit to first 100 for MVP)
                rows = []
                for row_idx, row in enumerate(
                    sheet.iter_rows(min_row=2, max_row=101, values_only=True)
                ):
                    if row_idx >= 100:  # Limit to 100 rows
                        break
                    # Convert row to list, handling None values
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    # Only add non-empty rows
                    if any(cell for cell in row_data):
                        rows.append(row_data)

                # Store sheet data
                ExcelData.objects.create(
                    upload=upload,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    row_data={"headers": headers, "rows": rows},
                )

            wb.close()

        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")

    def form_invalid(self, form):
        """Handle form validation errors."""
        if hasattr(self.request, "htmx") and self.request.htmx:
            return render(
                self.request,
                "excel_manager/partials/_upload_area.html",
                {"form": form, "error": True},
                status=400,
            )
        return super().form_invalid(form)


class ExcelDetailView(LoginRequiredMixin, DetailView):
    """View for displaying Excel file contents."""

    model = ExcelUpload
    template_name = "excel_manager/detail.html"
    context_object_name = "upload"

    def get_queryset(self):
        """Ensure users can only view their own uploads."""
        return ExcelUpload.objects.filter(user=self.request.user).prefetch_related(
            "sheets"
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Viewing: {self.object.original_filename}"

        # Add settings for AI feature check
        context["settings"] = settings

        # Check if there's a recent validation (within 1 hour)
        from datetime import timedelta
        recent_validation = self.object.ai_validations.filter(
            validated_at__gte=timezone.now() - timedelta(hours=1)
        ).first()
        context["recent_validation"] = recent_validation
        context["has_cached_validation"] = recent_validation is not None

        # Get the requested sheet or default to first
        sheet_index = self.request.GET.get("sheet", 0)
        try:
            sheet_index = int(sheet_index)
        except (ValueError, TypeError):
            sheet_index = 0

        # Get all sheets for tabs
        sheets = list(self.object.sheets.all())
        context["sheets"] = sheets

        # Get current sheet data
        if sheets:
            current_sheet = sheets[min(sheet_index, len(sheets) - 1)]
            context["current_sheet"] = current_sheet
            context["current_sheet_index"] = current_sheet.sheet_index
        else:
            context["current_sheet"] = None
            context["current_sheet_index"] = 0

        return context


def sheet_data_partial(request, pk, sheet_index):
    """HTMX partial for switching between sheets."""
    if not request.user.is_authenticated:
        return render(request, "403.html", status=403)

    upload = get_object_or_404(ExcelUpload, pk=pk, user=request.user)
    sheet = get_object_or_404(ExcelData, upload=upload, sheet_index=sheet_index)

    context = {
        "upload": upload,
        "current_sheet": sheet,
        "current_sheet_index": sheet_index,
    }

    return render(request, "excel_manager/partials/_data_table.html", context)


# AI Validation Constants
VALIDATION_SYSTEM_PROMPT = """You are a data quality analyst. Analyze Excel data and return ONLY valid JSON without any markdown formatting or code blocks.

Return exactly this structure (no markdown, no ```json blocks, just raw JSON):
{
  "valid_rows": integer,
  "warning_rows": integer,
  "error_rows": integer,
  "issues": [
    {"row": integer, "column": string, "issue": string, "severity": "error"|"warning"}
  ],
  "summary": string (2-3 sentences),
  "suggestions": [string],
  "severity": "low"|"medium"|"high"
}

Focus on: missing values, format inconsistencies, data type errors, duplicates, logical errors.
IMPORTANT: Return ONLY the JSON object, no explanations, no markdown."""


def format_validation_prompt(data):
    """Format Excel data for validation, optimizing token usage."""
    columns = data.get("columns", [])
    rows = data.get("sample", [])
    total_rows = data.get("total_rows", 0)

    # Format data for AI
    formatted_rows = []
    for i, row in enumerate(rows[:50], 1):  # Limit to 50 rows for token optimization
        row_dict = {col: val for col, val in zip(columns, row)}
        formatted_rows.append(f"Row {i}: {row_dict}")

    prompt = f"""Validate this Excel data:

Sheet: {data.get('sheet_name', 'Sheet1')}
Total rows in file: {total_rows}
Columns: {', '.join(columns)}

Sample data (first 50 rows):
{chr(10).join(formatted_rows)}

Analyze for data quality issues and return JSON as specified."""

    return prompt


def validate_excel_with_ai(excel_upload):
    """Validate Excel data using AI service."""
    start_time = time.time()

    # Initialize AI service
    service = AIService()

    # Prepare data sample
    data_sample = excel_upload.get_preview_data(rows=100)
    if not data_sample.get("sample"):
        raise ValueError("No data to validate")

    prompt = format_validation_prompt(data_sample)

    # Send to AI for validation
    result = service.send_message(prompt=prompt, system=VALIDATION_SYSTEM_PROMPT)

    if not result.get("success"):
        raise Exception(f"AI validation failed: {result.get('error', 'Unknown error')}")

    # Parse AI response
    content = result["content"]

    # Remove markdown code blocks if present
    if "```json" in content:
        content = content.split("```json")[1].split("```")[0].strip()
    elif "```" in content:
        content = content.split("```")[1].split("```")[0].strip()

    try:
        validation_result = json.loads(content)
    except json.JSONDecodeError:
        # Try to extract JSON from the content
        import re
        json_match = re.search(r'\{[^{}]*"valid_rows"[^{}]*\}', result["content"], re.DOTALL)
        if json_match:
            try:
                validation_result = json.loads(json_match.group())
            except:
                # If AI didn't return valid JSON, create a basic structure
                validation_result = {
                    "valid_rows": 0,
                    "warning_rows": 0,
                    "error_rows": 0,
                    "issues": [],
                    "summary": result["content"][:500],
                    "suggestions": [],
                    "severity": "unknown",
                }
        else:
            validation_result = {
                "valid_rows": 0,
                "warning_rows": 0,
                "error_rows": 0,
                "issues": [],
                "summary": result["content"][:500],
                "suggestions": [],
                "severity": "unknown",
            }

    # Calculate response time
    response_time_ms = int((time.time() - start_time) * 1000)

    # Extract suggestions if present
    suggestions_text = "\n".join(validation_result.get("suggestions", []))

    # Create validation record
    validation = AIValidation.objects.create(
        excel_upload=excel_upload,
        validation_result=validation_result,
        issues_found=len(validation_result.get("issues", [])),
        suggestions=suggestions_text,
        ai_metadata={
            "tokens": result.get("usage", {}),
            "model": result.get("model", settings.AI_CONFIG.get("MODEL")),
            "response_time_ms": response_time_ms,
        },
    )

    return validation


class ValidateWithAIView(LoginRequiredMixin, View):
    """HTMX endpoint for AI validation."""

    def post(self, request, pk):
        """Handle AI validation request."""
        import logging
        logger = logging.getLogger(__name__)

        logger.info(f"ValidateWithAIView called for pk={pk}")
        logger.info(f"POST data: {request.POST}")
        logger.info(f"force_refresh: {request.POST.get('force_refresh')}")

        # Check if AI features are enabled
        if not settings.AI_CONFIG.get("ENABLED", False):
            logger.error("AI features are disabled")
            return render(
                request,
                "excel_manager/partials/_ai_validation_error.html",
                {"error": "AI features are currently disabled"},
                status=503,
            )

        # Get the upload and verify ownership
        excel_upload = get_object_or_404(ExcelUpload, pk=pk, user=request.user)

        # Check if user wants to force a fresh validation
        force_refresh = request.POST.get("force_refresh", "false") == "true"
        logger.info(f"Force refresh: {force_refresh}")

        # Check for recent cached validation (within 1 hour)
        if not force_refresh:
            recent_validation = excel_upload.ai_validations.filter(
                validated_at__gte=timezone.now() - timedelta(hours=1)
            ).first()
        else:
            recent_validation = None

        if recent_validation:
            # Return cached result
            logger.info(f"Returning cached validation from {recent_validation.validated_at}")
            return render(
                request,
                "excel_manager/partials/_ai_validation_result.html",
                {
                    "validation": recent_validation,
                    "cached": True,
                    "upload": excel_upload,
                },
            )

        # Perform new validation
        logger.info("Performing fresh validation")
        try:
            validation = validate_excel_with_ai(excel_upload)
            logger.info(f"Fresh validation completed: {validation.id}")
            return render(
                request,
                "excel_manager/partials/_ai_validation_result.html",
                {"validation": validation, "cached": False, "upload": excel_upload},
            )
        except Exception as e:
            logger.error(f"Validation failed: {str(e)}")
            return render(
                request,
                "excel_manager/partials/_ai_validation_error.html",
                {"error": str(e)},
                status=500,
            )


class DeleteExcelView(LoginRequiredMixin, View):
    """HTMX endpoint for deleting Excel uploads."""

    def post(self, request, pk):
        """Handle delete request."""
        excel_upload = get_object_or_404(ExcelUpload, pk=pk, user=request.user)

        # Delete the file from storage if it exists
        if excel_upload.file:
            excel_upload.file.delete()

        # Delete the database record (cascade will handle related data)
        excel_upload.delete()

        # Return updated file list for HTMX
        if hasattr(request, "htmx") and request.htmx:
            uploads = ExcelUpload.objects.filter(
                user=request.user
            ).select_related("user")[:10]
            return render(
                request,
                "excel_manager/partials/_file_list.html",
                {"uploads": uploads, "success_message": "File deleted successfully"}
            )

        messages.success(request, "Excel file deleted successfully")
        return redirect("excel_manager:index")
