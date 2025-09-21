import io
import pytest
from openpyxl import Workbook
from django.core.files.uploadedfile import SimpleUploadedFile
from apps.excel_manager.models import ExcelUpload, ExcelData
from apps.users.models import User


@pytest.fixture
def excel_file():
    """Create a valid Excel file for testing."""
    # Create Excel file in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    headers = ["Name", "Email", "Age", "Department"]
    ws.append(headers)

    # Add some data
    data = [
        ["John Doe", "john@example.com", 30, "Engineering"],
        ["Jane Smith", "jane@example.com", 25, "Marketing"],
        ["Bob Johnson", "bob@example.com", 35, "Sales"],
    ]
    for row in data:
        ws.append(row)

    # Add second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Product", "Price"])
    ws2.append([1, "Widget", 19.99])
    ws2.append([2, "Gadget", 29.99])

    # Save to bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return SimpleUploadedFile(
        "test_data.xlsx",
        excel_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def large_excel_file():
    """Create a large Excel file (>5MB) for testing size limits."""
    wb = Workbook()
    ws = wb.active

    # Add headers
    headers = ["Col" + str(i) for i in range(100)]
    ws.append(headers)

    # Add lots of data to make file large (>5MB)
    for i in range(20000):
        row = ["X" * 50 for _ in range(100)]  # Longer strings to ensure size
        ws.append(row)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Ensure file is actually larger than 5MB
    content = excel_buffer.getvalue()
    actual_size = len(content)
    if actual_size < 5 * 1024 * 1024:
        # If still not big enough, pad with more data
        # This is a bit hacky but ensures we test the limit
        padding_needed = (5 * 1024 * 1024) - actual_size + 1024
        padded_content = content + b"X" * padding_needed
        return SimpleUploadedFile(
            "large_file.xlsx",
            padded_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return SimpleUploadedFile(
        "large_file.xlsx",
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def pdf_file_as_xlsx():
    """Create a PDF file with .xlsx extension for testing validation."""
    pdf_content = b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog >>\nendobj\n"
    return SimpleUploadedFile(
        "fake.xlsx",
        pdf_content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def authenticated_client(client, user):
    """Return a client with an authenticated user."""
    client.force_login(user)
    return client


@pytest.fixture
def other_user():
    """Create another user for testing access control."""
    return User.objects.create_user(email="other@example.com", password="password123")


@pytest.fixture
def excel_upload(user):
    """Create an ExcelUpload instance."""
    return ExcelUpload.objects.create(
        user=user,
        file="test.xlsx",
        original_filename="test.xlsx",
        file_hash="testhash123",
        file_size=1024,
        status=ExcelUpload.STATUS_COMPLETED,
    )


@pytest.fixture
def excel_upload_factory(db):
    """Factory for creating ExcelUpload instances."""

    def create_excel_upload(**kwargs):
        defaults = {
            "file": "test.xlsx",
            "original_filename": "test.xlsx",
            "file_hash": f"hash_{kwargs.get('user', 'test')}",
            "file_size": 1024,
            "status": ExcelUpload.STATUS_COMPLETED,
        }
        defaults.update(kwargs)
        return ExcelUpload.objects.create(**defaults)

    return create_excel_upload


@pytest.fixture
def excel_upload_with_data(excel_upload):
    """Create an ExcelUpload with associated ExcelData."""
    ExcelData.objects.create(
        upload=excel_upload,
        sheet_name="Sheet1",
        sheet_index=0,
        row_data={
            "headers": ["Name", "Email", "Age"],
            "rows": [
                ["John Doe", "john@example.com", "30"],
                ["Jane Smith", "jane@example.com", "25"],
                ["Bob Johnson", "", "35"],  # Missing email
                ["Alice Brown", "alice@invalid", "28"],  # Invalid email
                ["Charlie Davis", "charlie@example.com", ""],  # Missing age
            ],
        },
    )
    return excel_upload
